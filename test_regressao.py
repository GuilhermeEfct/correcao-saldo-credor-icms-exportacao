# -*- coding: utf-8 -*-
"""
test_regressao.py — os DOZE testes de aceitação da ferramenta. Ferramenta LOCAL de
desenvolvimento: não vai para o HUB (como o test_app.py).

    python test_regressao.py                      # só a fixture sintética
    python test_regressao.py "C:\\caminho\\do\\sped"  # + o caso real de referência

A fixture sintética (pasta `fixtures/`) valida o caminho do CÁLCULO: R$ 62.000,00.
O caso real valida os dois caminhos de NÃO-CÁLCULO — e o resultado esperado nele é
NENHUMA correção. Os relatórios reais NÃO ficam no repositório; passe a pasta onde
eles estão. Sem o argumento, os testes do caso real são anunciados como não
executados (nunca silenciosamente omitidos).
"""

import glob
import os
import shutil
import sys
import tempfile

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import logica

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FIXTURES = os.path.join(BASE_DIR, "fixtures")

CNPJ_FIXTURE = "12345678000190"
RAZAO_FIXTURE = "Empresa Exemplo Indústria Ltda"
RAZAO_REAL = "Empresa do caso de referência"


def cnpj_do_arquivo(caminho):
    """CNPJ lido do próprio relatório.

    Nenhum identificador de cliente fica no repositório: quem tem os arquivos roda
    o teste; quem não tem, só vê os testes da fixture.
    """
    parte = logica.listar_partes([caminho])[0]
    with logica.abrir_linhas(parte) as fluxo:
        return logica.parse_apuracao_mensal(fluxo, caminho)["cnpj"]

_falhas = []
_ok = 0


def checar(condicao, titulo, detalhe=""):
    global _ok
    if condicao:
        _ok += 1
        print("  [OK]    %s" % titulo)
    else:
        _falhas.append(titulo + (" — " + detalhe if detalhe else ""))
        print("  [FALHOU] %s%s" % (titulo, ("  ->  " + detalhe) if detalhe else ""))


def quase(a, b, tol=0.005):
    return a is not None and abs(a - b) <= tol


def por_cnpj(res, sufixo):
    for e in res["estabelecimentos"]:
        if e["cnpj"].endswith(sufixo):
            return e
    return None


def gerar_excel_tmp(res, nome):
    destino = os.path.join(tempfile.gettempdir(), "regressao_cscie", nome)
    return logica.gerar_excel_auditavel(res, destino)


def sem_clausula_de_exportacao(est):
    """O desfecho 'sem saldo credor' não pode apresentar a exportação como o obstáculo.

    Não basta procurar a palavra "exportação": o próprio template a cita para
    DESCARTÁ-LA ("independentemente de haver ou não operações de exportação no
    período"). O que reprova é a frase acusatória do outro desfecho ou qualquer
    cláusula extra grudada no fim — então o teste exige a string EXATA do template.
    """
    contexto = ""
    if est["ultimo_mes_com_saldo"]:
        contexto = logica.TPL_CONTEXTO_ULTIMO_SALDO.format(
            mes=est["ultimo_mes_com_saldo"]["mes"],
            valor=logica.fmt_brl(est["ultimo_mes_com_saldo"]["valor"]))
    esperado = logica.TPL_SEM_SALDO_CREDOR.format(mes_ref=est["mes_ref"], contexto=contexto)
    acusatorias = ("não houve operação de exportação", "também não", "e além disso",
                   "nem houve exporta")
    return (est["explicacao"] == esperado
            and not any(t in est["explicacao"].lower() for t in acusatorias))


# =============================================================================
# Fixture sintética — o caminho do cálculo
# =============================================================================
def testar_fixture():
    print("\n=== FIXTURE SINTÉTICA (caminho do cálculo) " + "=" * 34)
    arquivos = sorted(glob.glob(os.path.join(FIXTURES, "*.csv")))
    if len(arquivos) != 4:
        print("  !! fixture não encontrada em %s — rode: python "
              "gerar_fixture_sintetica.py fixtures" % FIXTURES)
        _falhas.append("fixture ausente")
        return None

    res = logica.processar_saldo_credor(arquivos, CNPJ_FIXTURE, RAZAO_FIXTURE)
    a, b, c, d = (por_cnpj(res, "000190"), por_cnpj(res, "000271"),
                  por_cnpj(res, "000352"), por_cnpj(res, "000433"))

    checar(quase(res["totais"]["correcao"], 62000.00),
           "TOTAL da fixture = R$ 62.000,00", logica.fmt_brl(res["totais"]["correcao"]))
    checar(quase(a["correcao"], 50000.00) and quase(a["percentual"], 0.10),
           "A calcula R$ 50.000,00 com 10,00% (transferência FORA do denominador)",
           "%s / %s" % (logica.fmt_brl(a["correcao"]), logica.fmt_pct(a["percentual"])))
    checar(quase(a["faturamento"], 1000000.00),
           "A: faturamento do mês = R$ 1.000.000,00, não R$ 6.000.000,00",
           logica.fmt_brl(a["faturamento"]))
    checar(b["status"] == logica.STATUS_SEM_EXPORTACAO and b["correcao"] is None,
           "B: sem exportação no último mês, sem retroceder para AGO/2025", b["status"])
    checar(c["status"] == logica.STATUS_SEM_SALDO_CREDOR and c["correcao"] is None,
           "C: sem saldo credor no último mês", c["status"])
    checar(sem_clausula_de_exportacao(c),
           "C: curto-circuito — a explicação não culpa a exportação "
           "(ele exportou R$ 30.000,00 em DEZ/2025)", c["explicacao"])
    checar("SET/2025" in c["explicacao"] and "45.000,00" in c["explicacao"],
           "C: frase de contexto com SET/2025 e R$ 45.000,00")
    checar(d["mes_ref"] == "NOV/2025" and "NOV/2025" in d["explicacao"],
           "D: mês de referência por estabelecimento (NOV/2025)", d["mes_ref"])
    # o mesmo relatório sai com 'JAN/2021' e com 'jan/21'; com ano de 2 dígitos o
    # parser não achava mês nenhum e recusava o arquivo como se estivesse errado
    checar(logica.normalizar_mes("jan/21") == "JAN/2021"
           and logica.normalizar_mes("DEZ/2025") == "DEZ/2025"
           and logica.chave_mes("dez/25") == (2025, 12)
           and logica.normalizar_mes("total") == "",
           "rótulo de mês aceito nos dois formatos ('jan/21' e 'JAN/2021')",
           logica.normalizar_mes("jan/21"))
    checar(quase(d["correcao"], 12000.00) and quase(d["percentual"], 0.12),
           "D calcula R$ 12.000,00 com 12,00%", logica.fmt_brl(d["correcao"]))
    checar(not [av for av in res["avisos"] if "Seção 19 divergiu" in av],
           "check Seção 1 × Seção 19 fecha nos 4 arquivos", "; ".join(res["avisos"]))
    return res


# =============================================================================
# Caso real de referência — os dois caminhos de NÃO-CÁLCULO
# =============================================================================
ESPERADO_REAL = {
    "000102": ("RS", "MAI/2026", 4343256.97, 0.0, 1814875.92, logica.STATUS_SEM_EXPORTACAO),
    "000374": ("ES", "MAI/2026", 1332773.85, 0.0, 18527.16, logica.STATUS_SEM_EXPORTACAO),
    "000455": ("ES", "MAI/2026", 843668.47, 0.0, 1343.85, logica.STATUS_SEM_EXPORTACAO),
    "000536": ("RS", "MAI/2026", 13400344.83, 0.0, 0.0, logica.STATUS_SEM_SALDO_CREDOR),
    "000617": ("SP", "ABR/2026", 2992360.75, 0.0, 921945.17, logica.STATUS_SEM_EXPORTACAO),
    "000706": ("RS", "MAI/2026", 0.0, 0.0, 56935.58, logica.STATUS_SEM_EXPORTACAO),
    "000889": ("RS", "MAI/2026", 34233.00, 0.0, 1072174.23, logica.STATUS_SEM_EXPORTACAO),
}
# valores que denunciam metodologia errada (seção 6.2 do briefing)
PROIBIDOS = {"retrocesso na série": 40897.08, "soma dos meses com exportação": 1610515.70,
             "percentual do período": 8240.72}


def testar_caso_real(pasta):
    print("\n=== CASO REAL DE REFERÊNCIA (não-cálculo) " + "=" * 35)
    arquivos = sorted(glob.glob(os.path.join(pasta, "ICMSProprio*.csv")))
    if len(arquivos) != 7:
        print("  -- NÃO EXECUTADO: esperava 7 relatórios ICMSProprio*.csv em %s "
              "(achei %d)." % (pasta, len(arquivos)))
        return None

    cnpj_real = cnpj_do_arquivo(arquivos[0])
    res = logica.processar_saldo_credor(arquivos, cnpj_real, RAZAO_REAL)

    # 1 — não inventar resultado
    checar(res["totais"]["correcao"] is None
           and all(e["correcao"] is None for e in res["estabelecimentos"])
           and all(e["percentual"] is None for e in res["estabelecimentos"]),
           "1. Nenhuma correção calculada — nem R$ 0,00 em coluna de valor",
           str(res["totais"]["correcao"]))

    # 2, 3, 4 — números que denunciariam metodologia errada
    todos = [e["correcao"] for e in res["estabelecimentos"] if e["correcao"] is not None]
    todos.append(res["totais"]["correcao"] or 0.0)
    for nome, proibido in PROIBIDOS.items():
        checar(not any(quase(v, proibido, 1.0) for v in todos),
               "%s. Não aparece R$ %s (%s)"
               % (2 + list(PROIBIDOS).index(nome), "{:,.2f}".format(proibido), nome))

    # 5, 7, 12 — mês, UF, valores e status por estabelecimento
    for sufixo, (uf, mes, fat, exp, saldo, status) in ESPERADO_REAL.items():
        e = por_cnpj(res, sufixo)
        okz = (e and e["uf"] == uf and e["mes_ref"] == mes and quase(e["faturamento"], fat)
               and quase(e["exportacao"], exp) and quase(e["saldo_credor"], saldo)
               and e["status"] == status)
        checar(okz, "…%s: %s, %s, %s, saldo %s, %s"
               % (sufixo, uf, mes, logica.fmt_brl(fat), logica.fmt_brl(saldo), status),
               "" if okz else "veio %s" % ({k: e[k] for k in
                                            ("uf", "mes_ref", "faturamento", "exportacao",
                                             "saldo_credor", "status")} if e else None))

    # 6 — curto-circuito: o sem-saldo não pode citar exportação
    e536 = por_cnpj(res, "000536")
    checar(sem_clausula_de_exportacao(e536),
           "6. Curto-circuito: explicação do …0536 não culpa a exportação",
           e536["explicacao"])
    checar(e536["ultimo_mes_com_saldo"]
           and e536["ultimo_mes_com_saldo"]["mes"] == "FEV/2026"
           and "122.907,55" in e536["explicacao"],
           "6b. …0536 traz o contexto FEV/2026 com R$ 122.907,55")

    # 7 — o último mês é por estabelecimento
    e617 = por_cnpj(res, "000617")
    checar("ABR/2026" in e617["explicacao"] and "MAI/2026" not in e617["explicacao"],
           "7. Explicação do …0617 cita ABR/2026 e não MAI/2026", e617["explicacao"])

    # 9 — check Seção 1 × Seção 19
    checar(not [av for av in res["avisos"] if "Seção 19 divergiu" in av],
           "9. Seção 1 × Seção 19 fecha nos 7 arquivos (R$ 0,00)",
           "; ".join(res["avisos"]))

    # janela de 60 meses sobre a série de 65 do caso real
    checar(res["periodo"]["coberto"] == "JUN/2021 a MAI/2026 (60 meses)",
           "Janela recorta os 65 meses do arquivo para os 60 não prescritos",
           res["periodo"]["coberto"])
    meses_export = {o["mes"] for e in res["estabelecimentos"] for o in e["exportacoes"]}
    checar("ABR/2021" not in meses_export,
           "Exportação de ABR/2021 (prescrita) não aparece nem como prova",
           str(sorted(meses_export)))

    # 10 — validação de CNPJ raiz: os 7 passam juntos; raiz estranha aborta
    checar(len(res["estabelecimentos"]) == 7,
           "10a. Os 7 estabelecimentos da mesma raiz passam juntos")
    intruso = os.path.join(FIXTURES, "FIXTURE_ICMSProprio_A_calcula.csv")
    try:
        logica.processar_saldo_credor(arquivos + [intruso], cnpj_real, RAZAO_REAL)
        checar(False, "10b. Arquivo de outra raiz aborta com erro de negócio",
               "processou sem reclamar")
    except logica.ErroDeNegocio as erro:
        checar("não corresponde à empresa consultada" in str(erro),
               "10b. Arquivo de outra raiz aborta com erro de negócio", str(erro)[:90])

    # 12 — saldo estático qualifica o primeiro critério
    e374 = por_cnpj(res, "000374")
    estaticos = sum(1 for r in e374["serie"] if quase(r["saldo_credor"] or 0, 18527.16))
    checar(e374["status"] == logica.STATUS_SEM_EXPORTACAO and estaticos >= 6,
           "12. Saldo estático (…0374, 6 meses iguais) passa o 1º critério e "
           "trava só na exportação", "%d meses iguais, status %s"
           % (estaticos, e374["status"]))
    return res


# =============================================================================
# 8 e 11 — a explicação é a MESMA string em todo canal; sem-oportunidade tem Excel
# =============================================================================
def testar_excel(res, nome_arquivo, titulo):
    from openpyxl import load_workbook
    caminho = gerar_excel_tmp(res, nome_arquivo)
    checar(os.path.getsize(caminho) > 5000,
           "11. Excel gerado e disponível (%s)" % titulo,
           "%d bytes" % os.path.getsize(caminho))

    wb = load_workbook(caminho)
    checar(wb.sheetnames == ["Cálculo", "Série Mensal (conferência)", "Exportações"],
           "Excel com as 3 abas na ordem", str(wb.sheetnames))
    ws = wb["Cálculo"]

    iguais = True
    for i, est in enumerate(res["estabelecimentos"]):
        if ws.cell(row=15 + i, column=11).value != est["explicacao"]:
            iguais = False
    checar(iguais, "8. Coluna Explicação do Excel == explicação do JSON, "
                   "palavra por palavra (%s)" % titulo)

    calculou = res["totais"]["correcao"] is not None
    for i, est in enumerate(res["estabelecimentos"]):
        r = 15 + i
        pct, corr = ws.cell(row=r, column=7).value, ws.cell(row=r, column=10).value
        esperado_pct = '=IF($I{0}<>"Calculado","",IF(E{0}=0,0,F{0}/E{0}))'.format(r)
        esperado_corr = '=IF($I{0}<>"Calculado","",H{0}*G{0})'.format(r)
        if pct != esperado_pct or corr != esperado_corr:
            checar(False, "Fórmulas reais na linha %d" % r, "%s | %s" % (pct, corr))
            break
    else:
        checar(True, "Fórmulas reais em % Exportação e Correção (nunca número solto)")

    n = len(res["estabelecimentos"])
    total_cel = ws.cell(row=15 + n, column=10).value
    checar(total_cel == '=IF(COUNT(J15:J{0})=0,"",SUM(J15:J{0}))'.format(14 + n),
           "TOTAL soma apenas os calculados", str(total_cel))
    veredito = ws["B11"].value
    checar((veredito == "SEM OPORTUNIDADE") if not calculou
           else quase(veredito, round(res["totais"]["correcao"], 2)),
           "Bloco de veredito coerente com o desfecho (%s)" % titulo, str(veredito))
    checar(ws["E11"].value == res["explicacao_consolidada"],
           "Explicação consolidada no Excel == a da tela")
    return caminho


# =============================================================================
# Janela de 60 meses — o prazo não prescrito
# =============================================================================
def testar_janela():
    print("\n=== JANELA DE %d MESES (prescrição) " % logica.MESES_ANALISE + "=" * 38)
    from datetime import date
    import gerar_fixture_sintetica as g

    # contagem de CALENDÁRIO, não de colunas presentes
    janela = logica.meses_da_janela(["JAN/2021", "MAI/2026", "JUN/2021", "DEZ/2025"])
    checar("JAN/2021" not in janela and "JUN/2021" in janela and "DEZ/2025" in janela,
           "JAN/2021 fica fora da janela ancorada em MAI/2026; JUN/2021 fica dentro",
           str(sorted(janela)))

    # estabelecimento inteiro fora da janela sai da análise, mas COM aviso
    antigo = dict(g.ESTABELECIMENTOS[0])
    antigo.update(cnpj="12345678000514", uf="PR",
                  meses=["JUL/2019", "AGO/2019", "SET/2019", "OUT/2019", "NOV/2019",
                         "DEZ/2019"])
    pasta = os.path.join(tempfile.gettempdir(), "regressao_cscie_janela")
    os.makedirs(pasta, exist_ok=True)
    prescrito = os.path.join(pasta, "ICMSProprio_PRESCRITO.csv")
    with open(prescrito, "w", encoding="utf-8-sig", newline="") as fh:
        fh.write(g.montar(antigo))

    arquivos = sorted(glob.glob(os.path.join(FIXTURES, "*.csv"))) + [prescrito]
    res = logica.processar_saldo_credor(arquivos, CNPJ_FIXTURE, RAZAO_FIXTURE)
    checar(len(res["estabelecimentos"]) == 4
           and not por_cnpj(res, "000514"),
           "Estabelecimento sem escrituração na janela fica fora da análise",
           str([e["cnpj_fmt"] for e in res["estabelecimentos"]]))
    checar(any("já está prescrito" in a for a in res["avisos"]),
           "…e a exclusão vira AVISO, não omissão silenciosa", "; ".join(res["avisos"]))
    checar(quase(res["totais"]["correcao"], 62000.00),
           "O total dos demais segue R$ 62.000,00")
    checar("6 meses" in res["periodo"]["coberto"],
           "O período exibido reflete só o que entrou na análise",
           res["periodo"]["coberto"])

    # a janela é contada do mês de corte (relatório reproduzível), e extração velha avisa
    checar(logica.aviso_extracao_antiga("MAI/2026", hoje=date(2026, 7, 31)) is None,
           "Extração recente não gera aviso de defasagem")
    checar(logica.aviso_extracao_antiga("DEZ/2019", hoje=date(2026, 7, 31)) is not None,
           "Extração antiga gera aviso de defasagem (sem alterar número algum)")
    shutil.rmtree(pasta, ignore_errors=True)


# =============================================================================
# Gate de permissão — nega por padrão, concede só quando o dict concede
# =============================================================================
def testar_gate_permissao():
    print("\n=== GATE DE PERMISSÃO (nega por padrão) " + "=" * 35)
    from flask import Flask
    import routes

    def status_com(ctx, provider=True):
        app = Flask(__name__)
        routes.init_app(app)
        routes.set_auth_provider((lambda _r: (True, ctx)) if provider else None)
        app.register_blueprint(routes.bp)
        with app.test_client() as c:
            return c.get("/tools/%s/limites" % routes.TOOL_ID).status_code

    # A forma anterior (`if isinstance(p, dict) and not p.get(...)`) tinha fail-open:
    # contexto sem `permissions`, ou com tipo inesperado, pulava o if e deixava entrar.
    for titulo, ctx in (
        ("contexto sem a chave permissions", {"login": "x"}),
        ("permissions = None", {"login": "x", "permissions": None}),
        ("permissions com tipo inesperado", {"login": "x", "permissions": ["icms"]}),
        ("permissions sem a chave do grupo", {"login": "x", "permissions": {"outra": True}}),
        ("permissions com a chave em False", {"login": "x", "permissions": {"icms": False}}),
    ):
        checar(status_com(ctx) == 403, "403 quando %s" % titulo, "veio %d" % status_com(ctx))

    checar(status_com({"login": "x", "permissions": {"icms": True}}) == 200,
           "200 quando a permissão do grupo concede")
    checar(status_com({"login": "x", "is_admin": True, "permissions": {}}) == 200,
           "200 para is_admin sem a chave do grupo")
    checar(status_com({}, provider=False) == 401,
           "401 sem provider registrado (o módulo não libera por conta própria)")
    checar(routes.PERMISSAO == "icms",
           "a chave é a do GRUPO (icms), não uma chave própria da ferramenta",
           routes.PERMISSAO)


# =============================================================================
# Mensagens de erro que se explicam
#
# O princípio da ferramenta é o analista SABER o que está ocorrendo. Um erro que só
# diz "confira se é o relatório certo" joga a suspeita no arquivo dele e o manda
# conferir à mão — que é o trabalho que a ferramenta deveria eliminar.
# =============================================================================
def testar_mensagens_de_erro():
    print("\n=== MENSAGENS DE ERRO SE EXPLICAM " + "=" * 42)
    pasta = os.path.join(tempfile.gettempdir(), "regressao_cscie_msg")
    os.makedirs(pasta, exist_ok=True)

    def erro_de(nome, conteudo):
        caminho = os.path.join(pasta, nome)
        with open(caminho, "w", encoding="utf-8-sig", newline="") as fh:
            fh.write(conteudo)
        try:
            logica.processar_saldo_credor([caminho], CNPJ_FIXTURE, RAZAO_FIXTURE)
            return ""
        except logica.ErroDeNegocio as e:
            return str(e)

    # variação de formato: diz o que achou, o que esperava, e NÃO culpa o arquivo
    msg = erro_de("ICMSProprio_mes_numerico.csv",
                  "ICMSProprio - 1. Resumo ICMS - x\r\n"
                  "DESCRIÇÃO;01/2021;02/2021\r\n"
                  "Valor Operacional - Saídas/Prestações;100,00;200,00\r\n")
    checar("01/2021" in msg and "JAN/2021" in msg and "jan/21" in msg
           and "não problema no seu arquivo" in msg,
           "formato de mês desconhecido: mostra as colunas achadas e os formatos aceitos",
           msg[:90])
    checar("seções 1" in msg,
           "…e diz quais seções do relatório foram reconhecidas", msg[:90])

    # truncado: seções existem, cabeçalho não
    msg = erro_de("ICMSProprio_truncado.csv",
                  "ICMSProprio - 1. Resumo ICMS - x\r\nICMSProprio - 19. Saídas - x\r\n"
                  "5101 - Venda;100,00\r\n")
    checar("truncado" in msg, "arquivo sem cabeçalho: aponta truncamento", msg[:90])

    # não é apuração, mas o nome engana
    msg = erro_de("ICMSProprio_outra_coisa.csv", "Data;Documento;Valor\r\n01/01/2026;1;2\r\n")
    checar("não é o relatório de Apuração de ICMS" in msg,
           "arquivo que não é apuração: diz que não achou seção alguma", msg[:90])

    shutil.rmtree(pasta, ignore_errors=True)


# =============================================================================
# Grupo 7 não é sinônimo de exportação
# =============================================================================
def testar_cfop_exportacao():
    print("\n=== CFOP DE EXPORTAÇÃO (grupo 7 não basta) " + "=" * 33)
    checar(logica.eh_export_direta("7101") and logica.eh_export_direta("7501"),
           "7101 e 7501 contam como exportação")
    for cfop, oque in (("7202", "devolução de compra"), ("7949", "saída não especificada"),
                       ("7206", "anulação de valor"), ("7930", "lançamento")):
        checar(not logica.eh_export_direta(cfop)
               and logica.eh_saida_exterior_nao_qualificada(cfop),
               "%s (%s) NÃO conta como exportação" % (cfop, oque))

    # o caso real: unicas saidas do mes eram 7202 e 7949 -> a regra "todo o grupo 7"
    # dava 100% de exportacao e devolvia o saldo credor INTEIRO
    pasta = os.path.join(tempfile.gettempdir(), "regressao_cscie_cfop")
    os.makedirs(pasta, exist_ok=True)
    caminho = os.path.join(pasta, "ICMSProprio_exterior_nao_qualificado.csv")
    with open(caminho, "w", encoding="utf-8-sig", newline="") as fh:
        fh.write("ICMSProprio - 1. Resumo ICMS - x\r\n"
                 "DESCRIÇÃO;NOV/2025;DEZ/2025\r\n"
                 "Valor Operacional - Saídas/Prestações;0,00;600.000,00\r\n\r\n"
                 "ICMSProprio - 19. Valor Operacional por CFOP - Saídas/Prestações - \r\n"
                 "DESCRIÇÃO;NOV/2025;DEZ/2025\r\n"
                 "7202 - Devol compra p/comercial;0,00;323.396,16\r\n"
                 "7949 - Outra saída não especificada;0,00;276.603,84\r\n\r\n"
                 "ICMSProprio - 22. Abertura Saldo a Transportar por CNPJ - \r\n"
                 "DESCRIÇÃO;NOV/2025;DEZ/2025\r\n"
                 "12345678000190 - RS;50.000,00;56.383,79\r\n")
    res = logica.processar_saldo_credor([caminho], CNPJ_FIXTURE, RAZAO_FIXTURE)
    e = res["estabelecimentos"][0]
    checar(e["status"] == logica.STATUS_SEM_EXPORTACAO and e["correcao"] is None,
           "só 7202 e 7949 no mês NÃO viram 100% de exportação",
           "%s / %s" % (e["status"], e["correcao"]))
    checar("7202" in e["explicacao"] and "7949" in e["explicacao"]
           and "não geram o crédito" in e["explicacao"],
           "…e a explicação NOMEIA os CFOPs e diz por que não contam", e["explicacao"][-120:])
    shutil.rmtree(pasta, ignore_errors=True)


def main():
    res_fix = testar_fixture()
    testar_janela()
    testar_gate_permissao()
    testar_mensagens_de_erro()
    testar_cfop_exportacao()
    if res_fix:
        print("\n--- Excel da fixture " + "-" * 55)
        testar_excel(res_fix, "fixture.xlsx", "fixture, com cálculo")

    pasta_real = sys.argv[1] if len(sys.argv) > 1 else os.environ.get("CSCIE_SPED_REAL", "")
    if pasta_real:
        res_real = testar_caso_real(pasta_real)
        if res_real:
            print("\n--- Excel do caso real " + "-" * 53)
            testar_excel(res_real, "caso_real.xlsx", "caso real, sem oportunidade")
    else:
        print("\n=== CASO REAL DE REFERÊNCIA " + "=" * 49)
        print("  -- NÃO EXECUTADO: passe a pasta dos relatórios reais como argumento.")
        print("     Os relatórios do cliente não ficam no repositório, de propósito.")

    print("\n" + "=" * 76)
    print("%d verificações OK, %d falha(s)." % (_ok, len(_falhas)))
    for f in _falhas:
        print("  FALHOU: %s" % f)
    shutil.rmtree(os.path.join(tempfile.gettempdir(), "regressao_cscie"),
                  ignore_errors=True)
    return 1 if _falhas else 0


if __name__ == "__main__":
    sys.exit(main())
