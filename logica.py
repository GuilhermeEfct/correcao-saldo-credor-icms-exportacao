# -*- coding: utf-8 -*-
"""
logica.py — Lógica fiscal PURA da ferramenta
"Correção do Saldo Credor de ICMS proporcional às Exportações".

Sem Flask, sem estado global, sem caminho absoluto.

O QUE A FERRAMENTA FAZ
----------------------
Sobre o ÚLTIMO MÊS DE APURAÇÃO de cada estabelecimento — e só sobre ele — avalia
duas condições em cascata CURTO-CIRCUITADA e devolve um de três desfechos:

    1) saldo credor de ICMS = 0            -> SEM_SALDO_CREDOR (termina aqui)
    2) saldo credor > 0, exportação = 0    -> SEM_EXPORTACAO
    3) saldo credor > 0 e exportação > 0   -> CALCULADO
       correção = saldo credor × (faturamento de exportação ÷ faturamento total)

Os desfechos 1 e 2 são RESULTADO VÁLIDO, não erro: as funções deste módulo
retornam o dicionário normalmente e NÃO levantam exceção. `ErroDeNegocio` está
reservada a falha real (arquivo ilegível, seção ausente, CNPJ divergente).

Entrada: relatório de Apuração de ICMS ("ICMSProprio") em CSV, um arquivo por
estabelecimento, aceito também em .txt e .zip. Lido linha a linha, em fluxo.

------------------------------------------------------------------------------
INTERFACE PÚBLICA (é isto que o routes.py chama — mantenha as assinaturas):

    processar_saldo_credor(arquivos, cnpj, razao_social, progress=None) -> dict
    gerar_excel_auditavel(resultado, saida_path) -> str
    ErroDeNegocio
------------------------------------------------------------------------------
"""

from __future__ import annotations

import codecs
import contextlib
import io
import itertools
import os
import re
import zipfile
from collections import namedtuple


class ErroDeNegocio(Exception):
    """Falha real, com mensagem humana. NUNCA usada para 'não há oportunidade'."""


# =============================================================================
# CFOPs — constantes de classificação
# =============================================================================

# EXPORTAÇÃO DIRETA — saída para o exterior. Todo CFOP do grupo 7.
def eh_export_direta(cfop: str) -> bool:
    return cfop.startswith("7")


# EXPORTAÇÃO INDIRETA — remessa com fim específico de exportação.
CFOP_EXPORT_INDIRETA = {"5501", "5502", "5503", "5504", "5505",
                        "6501", "6502", "6503", "6504", "6505"}

# FATURAMENTO (venda) — grupos x1xx e x4xx de venda.
# Transferência entre estabelecimentos (5151/5152/5153/6151/6152/6153), remessa,
# bonificação e devolução NÃO entram: transferência não é faturamento, e incluí-la
# no denominador derrubaria o percentual sem justificativa técnica.
CFOP_VENDA = {"5101", "5102", "5103", "5104", "5105", "5106", "5109", "5110", "5111",
              "5112", "5113", "5114", "5115", "5116", "5117", "5118", "5119", "5120",
              "5122", "5123", "5401", "5402", "5403", "5405", "5551", "5552",
              "6101", "6102", "6103", "6104", "6105", "6106", "6107", "6108", "6109",
              "6110", "6111", "6112", "6113", "6114", "6115", "6116", "6117", "6118",
              "6119", "6120", "6122", "6123", "6401", "6402", "6403", "6404",
              "6551", "6552"}

# DEVOLUÇÃO / RETORNO DE EXPORTAÇÃO (entradas, Seção 20) — abatida do faturamento
# de exportação do mês, com piso em zero.
#   1503-1506 / 2503-2506 : retorno de remessa com fim específico de exportação
#   3201/3202/3211        : devolução de venda vinda do exterior (exportação direta)
# Sem o piso, um retorno maior que a exportação do mês produziria exportação
# negativa e percentual negativo na aba de conferência.
CFOP_DEVOLUCAO_EXPORT = {"1503", "1504", "1505", "1506",
                         "2503", "2504", "2505", "2506",
                         "3201", "3202", "3211"}


def eh_export(cfop: str) -> bool:
    return eh_export_direta(cfop) or cfop in CFOP_EXPORT_INDIRETA


# =============================================================================
# Status e rótulos
# =============================================================================
STATUS_CALCULADO = "CALCULADO"
STATUS_SEM_EXPORTACAO = "SEM_EXPORTACAO"
STATUS_SEM_SALDO_CREDOR = "SEM_SALDO_CREDOR"

# Rótulo curto: colore a tela, filtra a tabela e vai para a coluna Status do
# Excel. O rótulo "Calculado" é lido pelas fórmulas da planilha — não altere.
ROTULO_STATUS = {
    STATUS_CALCULADO: "Calculado",
    STATUS_SEM_EXPORTACAO: "Sem exportação no último mês",
    STATUS_SEM_SALDO_CREDOR: "Sem saldo credor no último mês",
}


# =============================================================================
# Redação — os textos que o analista copia e cola para o cliente
#
# Ficam AQUI, num só lugar, porque a mesma string vai para o JSON, para a tela e
# para a coluna Explicação do Excel. Duplicá-los no screen.js faria as três
# versões divergirem com o tempo.
# =============================================================================
TPL_CALCULADO = (
    "No último mês de apuração ({mes_ref}) a empresa apresenta saldo credor de ICMS "
    "acumulado de {saldo} e faturamento de exportação de {export}, equivalente a "
    "{percentual} do faturamento total de {faturamento}. Aplicada essa proporção ao "
    "saldo credor, a oportunidade é de {correcao}."
)

TPL_SEM_EXPORTACAO = (
    "Identifiquei saldo credor de ICMS acumulado de {saldo} no último mês de apuração "
    "({mes_ref}), porém não houve operação de exportação nesse mês. Como a correção é "
    "proporcional ao faturamento de exportação do próprio mês de referência, não há "
    "oportunidade a ser calculada."
)

# Ancorada no MÊS, não no estado da empresa: "não possui mais saldo credor"
# afirmaria esgotamento permanente, e isso pode ser falso — no caso de referência
# um estabelecimento zerou o saldo em FEV/2026 e voltou a acumular em MAR/2026.
TPL_SEM_SALDO_CREDOR = (
    "Identifiquei que no último mês de apuração ({mes_ref}) não há saldo credor de ICMS "
    "acumulado e, com isso, não há oportunidade a ser calculada. Sem saldo credor não "
    "existe base sobre a qual aplicar a proporção das exportações, independentemente de "
    "haver ou não operações de exportação no período.{contexto}"
)

# Contexto opcional do desfecho 1. É informação, NÃO convite a recalcular por
# outro mês: nenhum valor de correção acompanha esta frase.
TPL_CONTEXTO_ULTIMO_SALDO = (
    " O último mês em que houve saldo credor foi {mes}, com {valor}."
)

TPL_CONSOLIDADO_CALCULADO = (
    "Dos {n} estabelecimentos analisados, {k} {apresentou} saldo credor de ICMS acumulado "
    "e operação de exportação no último mês de apuração. Aplicada a proporção das "
    "exportações ao saldo credor de cada estabelecimento, a oportunidade total é de "
    "{total}. O detalhamento por estabelecimento consta na tabela abaixo."
)

TPL_CONSOLIDADO_SEM_OPORTUNIDADE = (
    "Dos {n} estabelecimentos analisados, nenhum apresentou, simultaneamente, saldo "
    "credor de ICMS acumulado e operação de exportação no último mês de apuração. "
    "{composicao} Não há oportunidade a ser calculada."
)

# ---- Ressalvas do rodapé do Excel (e da seção homônima do README) -----------
RESSALVA_MES = (
    "O cálculo é feito exclusivamente sobre o último mês de apuração de cada "
    "estabelecimento, sem retrocesso na série histórica. O mês utilizado consta na "
    'coluna "Último mês de apuração".'
)
RESSALVA_JANELA = (
    "A análise abrange os últimos {limite} meses contados do mês de corte da extração "
    "({corte}), prazo dentro do qual a recuperação de crédito não está prescrita. Meses "
    "anteriores foram omitidos de todo o relatório — inclusive da série de conferência e "
    "das operações de exportação —, e não integram nenhum total apresentado."
)
RESSALVA_NATUREZA_SALDO = (
    "O valor utilizado é o saldo credor a transportar apurado no registro E110 "
    "(VL_SLD_CREDOR_TRANSPORTAR), que corresponde ao saldo da apuração ordinária. Não "
    "equivale, por si, a crédito acumulado formalmente gerado e apropriado nos termos do "
    "art. 25, §1º, I, da LC 87/96 — cuja constituição depende do rito da legislação de "
    "cada estado (por exemplo, a sistemática de crédito acumulado do RICMS/SP ou regime "
    "especial no RS). O valor aqui apurado é a estimativa da parcela proporcional às "
    "exportações, e não um crédito já habilitado à transferência."
)
RESSALVA_FATURAMENTO = (
    "Considera exclusivamente CFOPs de venda somados aos de exportação. Transferências "
    "entre estabelecimentos da própria empresa (5151/5152/5153/6151/6152/6153), remessas, "
    "bonificações e devoluções não integram o denominador."
)
RESSALVA_EXPORT_INDIRETA = (
    "As operações classificadas nos CFOPs 5501/5502/6501/6502 são remessas com fim "
    "específico de exportação e foram computadas pelo valor de face. O direito ao crédito "
    "depende da efetiva exportação no prazo legal pelo destinatário, o que não é "
    "verificável a partir da apuração de ICMS."
)
RESSALVA_DEVOLUCAO = (
    "Foram identificadas, nas entradas, operações de devolução/retorno de exportação "
    "(CFOPs {cfops}), abatidas do faturamento de exportação do mês em que ocorreram. "
    "Quando o valor devolvido supera a exportação do próprio mês, o resultado é limitado "
    "a zero — nunca a um valor negativo."
)
MSG_RAIZES_DIFERENTES = (
    "Os relatórios enviados pertencem a empresas diferentes (raízes de CNPJ {raizes}). "
    "Cada análise trata de uma empresa: envie os relatórios de uma raiz por vez."
)

RESSALVA_ORIGEM = (
    "Seções 19 (Valor Operacional por CFOP - Saídas/Prestações) e 22 (Abertura Saldo a "
    "Transportar por CNPJ) do relatório de apuração de ICMS, correspondentes aos "
    "registros C190 e E110 da EFD ICMS/IPI. Período coberto: {periodo}. Mês de corte da "
    "extração: {corte}."
)


# =============================================================================
# Formatadores e ordenação de mês
# =============================================================================
MESES_PTBR = {m: i for i, m in enumerate(
    ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN",
     "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"], start=1)}

# JANELA DE ANÁLISE — 60 meses contados do mês de corte da extração.
#
# É o prazo dentro do qual a recuperação de crédito não está prescrita. Mês fora da
# janela NÃO APARECE em lugar nenhum: nem na série de conferência, nem na aba de
# exportações, nem na contagem do período. Não é "marcar em amarelo" — é omitir,
# porque dado prescrito exibido ao lado do apurável convida a somar os dois.
MESES_ANALISE = 60

# O mesmo relatório sai em dois formatos de rótulo, dependendo de como foi exportado:
# 'JAN/2021' e 'jan/21'. Aceitar os dois é obrigatório — com ano de 2 dígitos o
# parser não achava mês nenhum e o arquivo era recusado como se estivesse errado.
_RE_MES = re.compile(r"^([A-Z]{3})/(\d{2}|\d{4})$")
_RE_SECAO = re.compile(r"^ICMSProprio\s*-\s*(\d+)\s*[.\-]")
_RE_CFOP = re.compile(r"^(\d{4})\s*-\s*(.*)$")
_RE_CNPJ_UF = re.compile(r"^(\d{14})\s*-\s*([A-Za-z]{2})?")


def fmt_brl(v) -> str:
    """1814875.92 -> 'R$ 1.814.875,92'"""
    s = "{:,.2f}".format(float(v or 0.0))
    return "R$ " + s.replace(",", "\x00").replace(".", ",").replace("\x00", ".")


def fmt_pct(fracao) -> str:
    """0.1967 -> '19,67%'  (2 casas decimais, como manda a metodologia)"""
    return "{:.2f}%".format(float(fracao or 0.0) * 100).replace(".", ",")


def fmt_mes(rotulo: str) -> str:
    """Rótulo do mês em forma canônica ('MAI/2026')."""
    return normalizar_mes(rotulo) or (rotulo or "").strip().upper()


def normalizar_mes(rotulo: str) -> str:
    """'jan/21' e 'JAN/2021' -> 'JAN/2021'. Devolve '' se não for rótulo de mês.

    O ano de 2 dígitos vira 20XX: este relatório é de apuração de ICMS, e não existe
    escrituração de 19XX no escopo da ferramenta.
    """
    m = _RE_MES.match((rotulo or "").strip().upper())
    if not m or m.group(1) not in MESES_PTBR:
        return ""
    ano = m.group(2)
    return "%s/%s" % (m.group(1), ano if len(ano) == 4 else "20" + ano)


def fmt_cnpj(digitos: str) -> str:
    """'12345678000190' -> '12.345.678/0001-90'"""
    d = _so_digitos(digitos)
    if len(d) != 14:
        return digitos or ""
    return "%s.%s.%s/%s-%s" % (d[:2], d[2:5], d[5:8], d[8:12], d[12:])


def chave_mes(rotulo: str):
    """Ordenação por (ano, mês). Ordem alfabética destruiria a série."""
    canonico = normalizar_mes(rotulo)
    if not canonico:
        return (0, 0)
    sigla, ano = canonico.split("/")
    return (int(ano), MESES_PTBR[sigla])


def eh_rotulo_mes(texto: str) -> bool:
    return bool(normalizar_mes(texto))


def _indice_calendario(rotulo: str) -> int:
    ano, mes = chave_mes(rotulo)
    return ano * 12 + mes


def aviso_extracao_antiga(mes_corte: str, hoje=None, limite: int = MESES_ANALISE):
    """Avisa quando a extração é velha — e NÃO mexe em número nenhum.

    A janela é sempre contada do mês de corte, nunca da data de hoje: só assim o
    mesmo arquivo produz o mesmo relatório daqui a um ano, o que um documento
    auditável exige. O efeito colateral é que um SPED antigo passaria como se nada
    estivesse prescrito — daí o aviso, que é onde a data de hoje entra.
    """
    from datetime import date
    hoje = hoje or date.today()
    atraso = (hoje.year * 12 + hoje.month) - _indice_calendario(mes_corte)
    if atraso <= 12:
        return None
    return ("O mês de corte da extração é %s, há %d meses. A janela de %d meses é contada "
            "desse mês, então confira se o SPED está atualizado: com extração antiga, "
            "parte do período analisado pode já estar prescrita hoje."
            % (fmt_mes(mes_corte), atraso, limite))


def meses_da_janela(rotulos, limite: int = MESES_ANALISE) -> set:
    """Os `limite` meses NÃO PRESCRITOS, ancorados no mês mais recente do lote.

    A conta é de CALENDÁRIO, não de colunas presentes: se a série tiver buraco, pegar
    "as últimas 60 colunas" alcançaria mês já prescrito. A prescrição é fato do
    calendário. A âncora é o mês de corte da extração — a mesma para todos os
    estabelecimentos, porque o prazo não corre por inscrição estadual.
    """
    presentes = sorted({fmt_mes(m) for m in rotulos if eh_rotulo_mes(m)}, key=chave_mes)
    if not presentes:
        return set()
    corte = _indice_calendario(presentes[-1]) - (limite - 1)
    return {m for m in presentes if _indice_calendario(m) >= corte}


def _conc(n: int, singular: str, plural: str) -> str:
    return singular if n == 1 else plural


# =============================================================================
# Utilidades de campo
# =============================================================================
def _clean(s: str) -> str:
    """Normaliza campo: remove BOM, prefixo ="..." (Excel-texto) e aspas."""
    if s is None:
        return ""
    s = str(s).replace("﻿", "").strip()
    if s.startswith("="):
        s = s[1:]
    return s.strip().strip('"').strip()


def _num(s):
    """Número pt-BR ('1.234.567,89') -> float.

    Devolve None para célula VAZIA — e a diferença importa: mês sem escrituração
    do estabelecimento é ausência de dado, não saldo zero.
    """
    s = _clean(s).replace("R$", "").replace(" ", "").replace("\xa0", "")
    if s in ("", "-", "—", "--"):
        return None
    negativo = s.startswith("(") and s.endswith(")")
    if negativo:
        s = s[1:-1]
    if s.startswith("-"):
        negativo, s = True, s[1:]
    # ',' é o decimal; '.' é separador de milhar
    s = s.replace(".", "").replace(",", ".") if "," in s else s.replace(".", "")
    try:
        v = float(s)
    except ValueError:
        return None
    return -v if negativo else v


def _so_digitos(s: str) -> str:
    return "".join(ch for ch in (s or "") if ch.isdigit())


# =============================================================================
# Leitura EM FLUXO dos arquivos enviados (caminho no disco ou (nome, bytes)).
#
# Reaproveitado da ferramenta-irmã `credito-icms-uso-consumo-exportacao`, que lê
# o mesmo relatório: nada de carregar arquivo inteiro na memória — cada parte é
# lida linha a linha, inclusive membros de .zip, e o pico de RAM não acompanha o
# tamanho da entrada.
# =============================================================================
_Parte = namedtuple("_Parte", "nome origem membro")


def _membros_zip(origem):
    abrir = (origem[1] if origem[0] == "disco" else io.BytesIO(origem[1]))
    with zipfile.ZipFile(abrir) as zf:
        return [m for m in zf.namelist() if m.lower().endswith((".csv", ".txt"))]


def listar_partes(arquivos):
    """Descreve as partes a processar SEM ler conteúdo (só o índice do zip)."""
    partes = []
    for item in arquivos:
        if isinstance(item, (tuple, list)):
            nome, origem = item[0], ("memoria", item[1])
        else:
            nome, origem = item, ("disco", item)
        if str(nome).lower().endswith(".zip"):
            try:
                membros = _membros_zip(origem)
            except (zipfile.BadZipFile, OSError) as e:
                raise ErroDeNegocio(
                    "Não foi possível abrir o arquivo compactado %s (%s). Reenvie o .zip "
                    "ou mande os CSV soltos." % (os.path.basename(str(nome)), e)
                )
            for membro in membros:
                partes.append(_Parte(membro, origem, membro))
        else:
            partes.append(_Parte(nome, origem, None))
    return partes


def _encoding_do_fluxo(bruto) -> str:
    """Descobre a codificação olhando só o começo do fluxo, sem consumi-lo."""
    amostra = b""
    if hasattr(bruto, "peek"):
        amostra = bruto.peek(65536)[:65536]
    elif bruto.seekable():
        amostra = bruto.read(65536)
        bruto.seek(0)
    if amostra.startswith(codecs.BOM_UTF8):
        return "utf-8-sig"
    try:
        # incremental: não acusa erro se a amostra cortar um caractere no meio
        codecs.getincrementaldecoder("utf-8")().decode(amostra, False)
        return "utf-8"
    except UnicodeDecodeError:
        return "latin-1"


@contextlib.contextmanager
def abrir_linhas(parte):
    """Abre a parte e entrega um iterador de linhas de texto."""
    aberto = []
    try:
        if parte.membro is not None:
            zf = zipfile.ZipFile(parte.origem[1] if parte.origem[0] == "disco"
                                 else io.BytesIO(parte.origem[1]))
            aberto.append(zf)
            bruto = zf.open(parte.membro)
        elif parte.origem[0] == "disco":
            bruto = open(parte.origem[1], "rb")
        else:
            bruto = io.BytesIO(parte.origem[1])
        aberto.append(bruto)
        texto = io.TextIOWrapper(bruto, encoding=_encoding_do_fluxo(bruto),
                                 errors="replace", newline="")
        aberto.append(texto)
        yield texto
    finally:
        for f in reversed(aberto):
            try:
                f.close()
            except Exception:  # noqa: BLE001
                pass


def eh_relatorio_apuracao(nome: str, primeiras_linhas) -> bool:
    """Classifica pelo CONTEÚDO: o marcador 'ICMSProprio' está na 1ª linha."""
    amostra = "".join(primeiras_linhas)
    if "ICMSProprio" in amostra[:5000] or "Valor Operacional por CFOP" in amostra[:20000]:
        return True
    n = (nome or "").lower()
    return "icmsproprio" in n or "apura" in n


# =============================================================================
# Parser — Seções 1, 19, 20, 21 e 22, com valor POR MÊS, numa única passada
# =============================================================================
def parse_apuracao_mensal(linhas, nome_arquivo: str = "") -> dict:
    """
    Extrai de UM relatório de apuração, lendo linha a linha:

      cnpj / uf              : Seção 22, com fallback na 21 e nos 14 primeiros
                               dígitos do começo do arquivo
      meses                  : rótulos das colunas ('JAN/2021' ...). O número de
                               colunas varia por estabelecimento — filial aberta
                               depois tem menos meses. O alinhamento é sempre
                               pelo RÓTULO, nunca pela posição.
      cfops_saida            : {cfop: {mes: valor}}  Seção 19
      descricao_cfop         : {cfop: descrição}
      saldo_credor           : {mes: float|None}     Seção 22 (E110)
      devolucao_export       : {cfop: {mes: valor}}  Seção 20, só CFOPs de
                               devolução/retorno de exportação
      oper_saidas_s1         : {mes: float}          Seção 1, para o check 3.6b
    """
    meses: list[str] = []
    cnpj = uf = ""
    cfops_saida: dict[str, dict[str, float]] = {}
    descricao_cfop: dict[str, str] = {}
    saldo_credor: dict[str, float | None] = {}
    devolucao_export: dict[str, dict[str, float]] = {}
    oper_saidas_s1: dict[str, float] = {}
    inicio: list[str] = []
    secao = None

    # `colunas` guarda a posição de cada coluna do cabeçalho, com "" onde a coluna não
    # é mês (o relatório às vezes traz ';;;;' de preenchimento no fim da linha). O
    # alinhamento dos valores é por POSIÇÃO, então a lista posicional não pode ser
    # filtrada — quem é filtrada é `meses`, usada para percorrer a série.
    colunas: list[str] = []

    def _valores(vals, destino_por_mes):
        for i, bruto in enumerate(vals[:len(colunas)]):
            mes = colunas[i]
            if not mes:
                continue
            v = _num(bruto)
            if v is not None:
                destino_por_mes[mes] = destino_por_mes.get(mes, 0.0) + v

    for bruta in linhas:
        ln = bruta.rstrip("\r\n")
        if len(inicio) < 80:
            inicio.append(ln)

        m = _RE_SECAO.match(ln.lstrip("﻿"))
        if m:
            secao = m.group(1)
            continue
        if secao is None or not ln.strip():
            continue

        partes = ln.split(";")
        rotulo = _clean(partes[0])
        valores = partes[1:]

        # cabeçalho de meses: idêntico em todas as seções do arquivo
        if rotulo.upper().startswith("DESCRI"):
            if not colunas:
                candidatas = [normalizar_mes(_clean(c)) for c in valores]
                if any(candidatas):
                    colunas = candidatas
                    meses = [m for m in candidatas if m]
            continue
        if not colunas:
            continue

        if secao == "19":
            mc = _RE_CFOP.match(rotulo)
            if not mc:
                continue                      # linha fora do padrão 'NNNN - ': ignora
            cfop = mc.group(1)
            descricao_cfop.setdefault(cfop, mc.group(2).strip())
            _valores(valores, cfops_saida.setdefault(cfop, {}))

        elif secao == "20":
            mc = _RE_CFOP.match(rotulo)
            if not mc or mc.group(1) not in CFOP_DEVOLUCAO_EXPORT:
                continue
            cfop = mc.group(1)
            descricao_cfop.setdefault(cfop, mc.group(2).strip())
            _valores(valores, devolucao_export.setdefault(cfop, {}))

        elif secao in ("21", "22"):
            mc = _RE_CNPJ_UF.match(rotulo)
            if not mc:
                continue
            if not cnpj:
                cnpj = mc.group(1)
                uf = (mc.group(2) or "").upper()
            if secao == "22":
                # aqui o None PRECISA sobreviver: célula vazia = mês sem escrituração
                for i, bruto in enumerate(valores[:len(colunas)]):
                    if colunas[i]:
                        saldo_credor[colunas[i]] = _num(bruto)

        elif secao == "1" and rotulo.lower().startswith("valor operacional - sa"):
            _valores(valores, oper_saidas_s1)

    if not cnpj:
        # fallback: primeiros 14 dígitos que aparecerem no começo do arquivo
        m = re.search(r"(\d{14})", _so_digitos("".join(inicio)))
        cnpj = m.group(1) if m else ""

    if not meses:
        raise ErroDeNegocio(
            "Não foi possível identificar os períodos de apuração no arquivo %s. "
            "Confira se é o relatório de Apuração de ICMS (ICMSProprio) exportado em CSV."
            % (os.path.basename(nome_arquivo) or "enviado")
        )

    return {
        "arquivo": os.path.basename(nome_arquivo),
        "cnpj": cnpj,
        "uf": uf,
        "meses": meses,
        "cfops_saida": cfops_saida,
        "descricao_cfop": descricao_cfop,
        "saldo_credor": saldo_credor,
        "devolucao_export": devolucao_export,
        "oper_saidas_s1": oper_saidas_s1,
    }


# =============================================================================
# Série mensal, mês de referência e a cascata dos três desfechos
# =============================================================================
def serie_mensal(ap: dict) -> list[dict]:
    """Série ordenada do estabelecimento. Alimenta a aba de conferência.

    O CÁLCULO NÃO PERCORRE ESTA SÉRIE — ela serve para identificar o último mês
    e para o revisor conferir os insumos.
    """
    saida = []
    for mes in sorted(ap["meses"], key=chave_mes):
        exp_d = sum(v.get(mes, 0.0) for c, v in ap["cfops_saida"].items()
                    if eh_export_direta(c))
        exp_i = sum(v.get(mes, 0.0) for c, v in ap["cfops_saida"].items()
                    if c in CFOP_EXPORT_INDIRETA)
        devolucao = sum(v.get(mes, 0.0) for v in ap["devolucao_export"].values())
        exportacao = max(0.0, exp_d + exp_i - devolucao)
        venda = sum(v.get(mes, 0.0) for c, v in ap["cfops_saida"].items()
                    if c in CFOP_VENDA)
        saida.append({
            "mes": mes,
            "faturamento": venda + exportacao,
            "exportacao": exportacao,
            "exportacao_direta": exp_d,
            "exportacao_indireta": exp_i,
            "devolucao_export": devolucao,
            # o piso em zero foi acionado: a devolução do mês supera a exportação
            # do próprio mês (o retorno se refere a exportação de período anterior)
            "exportacao_limitada": bool(devolucao and devolucao > exp_d + exp_i),
            # None preservado: mês sem escrituração fica em branco, não em zero
            "saldo_credor": ap["saldo_credor"].get(mes),
        })
    return saida


def mes_referencia(ap: dict) -> str:
    """O mês MAIS RECENTE em que ESTE estabelecimento tem escrituração.

    É a última coluna de mês do arquivo dele — não o último mês do lote, e não o
    último mês com movimento (há estabelecimento cujo último mês tem faturamento
    zero e ainda assim é o mês de referência).
    """
    return sorted(ap["meses"], key=chave_mes)[-1]


def _linhas_exportacao(ap: dict, serie: list[dict]) -> list[dict]:
    """Uma linha por (CFOP de exportação × mês) com valor — prova documental.

    Restrita aos meses da série, isto é, à janela não prescrita: operação de mês
    prescrito não entra nem como prova.
    """
    na_janela = {r["mes"] for r in serie}
    linhas = []
    for cfop, por_mes in ap["cfops_saida"].items():
        if not eh_export(cfop):
            continue
        tipo = "Direta" if eh_export_direta(cfop) else "Indireta"
        for mes, valor in por_mes.items():
            if valor and mes in na_janela:
                linhas.append({"mes": mes, "cfop": cfop, "valor": valor, "tipo": tipo,
                               "descricao": ap["descricao_cfop"].get(cfop, "")})
    for cfop, por_mes in ap["devolucao_export"].items():
        for mes, valor in por_mes.items():
            if valor and mes in na_janela:
                linhas.append({"mes": mes, "cfop": cfop, "valor": -valor,
                               "tipo": "Devolução de exportação",
                               "descricao": ap["descricao_cfop"].get(cfop, "")})
    linhas.sort(key=lambda r: (chave_mes(r["mes"]), r["cfop"]))
    return linhas


def _ultimo_mes_com_saldo(serie: list[dict], antes_de: str):
    """(mês, valor) do último mês com saldo credor > 0 antes do mês de referência."""
    limite = chave_mes(antes_de)
    anteriores = [r for r in serie
                  if chave_mes(r["mes"]) < limite and (r["saldo_credor"] or 0.0) > 0]
    if not anteriores:
        return None
    ultimo = anteriores[-1]
    return (ultimo["mes"], ultimo["saldo_credor"])


def avaliar_estabelecimento(ap: dict) -> dict:
    """Aplica a cascata CURTO-CIRCUITADA e devolve o desfecho — sempre com sucesso."""
    serie = serie_mensal(ap)
    mes_ref = mes_referencia(ap)
    do_mes = next(r for r in serie if r["mes"] == mes_ref)

    saldo = do_mes["saldo_credor"] or 0.0
    faturamento = do_mes["faturamento"]
    exportacao = do_mes["exportacao"]
    contexto_saldo = _ultimo_mes_com_saldo(serie, mes_ref)

    if saldo <= 0:
        # TERMINAL. Não consulta a exportação e não a menciona: sem saldo credor
        # não existe base para proporção nenhuma, e citar a exportação sugeriria,
        # falsamente, que ela era o obstáculo.
        status = STATUS_SEM_SALDO_CREDOR
        percentual = correcao = None
        contexto = ""
        if contexto_saldo:
            contexto = TPL_CONTEXTO_ULTIMO_SALDO.format(
                mes=fmt_mes(contexto_saldo[0]), valor=fmt_brl(contexto_saldo[1]))
        explicacao = TPL_SEM_SALDO_CREDOR.format(mes_ref=fmt_mes(mes_ref),
                                                 contexto=contexto)
    elif exportacao <= 0:
        status = STATUS_SEM_EXPORTACAO
        percentual = correcao = None
        explicacao = TPL_SEM_EXPORTACAO.format(mes_ref=fmt_mes(mes_ref),
                                               saldo=fmt_brl(saldo))
    else:
        status = STATUS_CALCULADO
        # sem arredondar a proporção: o número da tela tem de ser o mesmo que a
        # fórmula viva do Excel produz ao recalcular
        percentual = exportacao / faturamento if faturamento else 0.0
        correcao = saldo * percentual
        explicacao = TPL_CALCULADO.format(
            mes_ref=fmt_mes(mes_ref), saldo=fmt_brl(saldo), export=fmt_brl(exportacao),
            percentual=fmt_pct(percentual), faturamento=fmt_brl(faturamento),
            correcao=fmt_brl(correcao))

    exportacoes = _linhas_exportacao(ap, serie)
    return {
        "arquivo": ap["arquivo"],
        "cnpj": ap["cnpj"],
        "cnpj_fmt": fmt_cnpj(ap["cnpj"]),
        "uf": ap["uf"],
        "mes_ref": fmt_mes(mes_ref),
        "faturamento": faturamento,
        "exportacao": exportacao,
        "exportacao_direta": do_mes["exportacao_direta"],
        "exportacao_indireta": do_mes["exportacao_indireta"],
        "saldo_credor": saldo,
        "percentual": percentual,
        "correcao": correcao,
        "status": status,
        "rotulo_status": ROTULO_STATUS[status],
        "explicacao": explicacao,
        "serie": serie,
        "exportacoes": exportacoes,
        "ultimo_mes_com_saldo": ({"mes": fmt_mes(contexto_saldo[0]),
                                  "valor": contexto_saldo[1]} if contexto_saldo else None),
        # derivadas das operações da JANELA, não das chaves do arquivo: ressalva sobre
        # operação prescrita seria nota a respeito de dado que o relatório não mostra
        "tem_export_indireta": any(o["tipo"] == "Indireta" for o in exportacoes),
        "cfops_devolucao": sorted({o["cfop"] for o in exportacoes
                                   if o["tipo"] == "Devolução de exportação"}),
        "meses_limitados": [r["mes"] for r in serie if r["exportacao_limitada"]],
        "primeiro_mes": fmt_mes(serie[0]["mes"]),
    }


def explicacao_consolidada(estabs: list[dict], total: float) -> str:
    """Uma ou duas frases para o lote. Com um só estabelecimento, é a dele."""
    if len(estabs) == 1:
        return estabs[0]["explicacao"]

    calculados = [e for e in estabs if e["status"] == STATUS_CALCULADO]
    n = len(estabs)
    if calculados:
        return TPL_CONSOLIDADO_CALCULADO.format(
            n=n, k=len(calculados),
            apresentou=_conc(len(calculados), "apresentou", "apresentaram"),
            total=fmt_brl(total))

    n_sem_export = sum(1 for e in estabs if e["status"] == STATUS_SEM_EXPORTACAO)
    n_sem_saldo = sum(1 for e in estabs if e["status"] == STATUS_SEM_SALDO_CREDOR)
    trechos = []
    if n_sem_export:
        trechos.append("%d %s saldo credor mas não %s exportação no mês" % (
            n_sem_export, _conc(n_sem_export, "possui", "possuem"),
            _conc(n_sem_export, "registrou", "registraram")))
    if n_sem_saldo:
        trechos.append("%d não %s saldo credor" % (
            n_sem_saldo, _conc(n_sem_saldo, "apresenta", "apresentam")))
    composicao = ("; ".join(trechos) + ".") if trechos else ""
    return TPL_CONSOLIDADO_SEM_OPORTUNIDADE.format(n=n, composicao=composicao)


# =============================================================================
# Validações de consistência (baratas e obrigatórias)
# =============================================================================
def validar_cnpj_raiz(estabs: list[dict], cnpj_tela: str, razao_social: str) -> None:
    """Raiz (8 primeiros dígitos) dos arquivos × CNPJ consultado na tela.

    Se divergir, ABORTA. O Excel sai com a razão social vinda da tela e o saldo
    credor vindo dos arquivos: se forem de empresas diferentes, o documento — que
    pode virar anexo processual — fica silenciosamente errado. Estabelecimentos
    diferentes da MESMA raiz são o caso normal e passam.
    """
    raiz_tela = _so_digitos(cnpj_tela)[:8]
    if len(raiz_tela) != 8:
        # cobre os dois caminhos: identificação automática que não achou o CNPJ nos
        # arquivos, e analista que digitou algo incompleto
        raise ErroDeNegocio(
            "Não foi possível identificar o CNPJ da empresa nos relatórios enviados. "
            "Informe o CNPJ no campo da tela e calcule novamente.")
    for e in estabs:
        raiz = _so_digitos(e["cnpj"])[:8]
        if raiz and raiz != raiz_tela:
            raise ErroDeNegocio(
                "O arquivo %s pertence ao CNPJ %s, que não corresponde à empresa "
                "consultada (%s — %s). Confira os arquivos enviados."
                % (e["arquivo"] or "enviado", fmt_cnpj(e["cnpj"]),
                   fmt_cnpj(cnpj_tela), razao_social or "razão social não informada"))


def conferir_secao1_x_secao19(ap: dict, tolerancia: float = 0.01) -> list[str]:
    """Seção 1 'Valor Operacional - Saídas' × soma de todos os CFOPs da Seção 19.

    Pega arquivo truncado ou exportação incompleta ANTES de a ferramenta produzir
    um percentual errado com aparência de certo. Avisa, não aborta.
    """
    avisos = []
    if not ap["oper_saidas_s1"]:
        return avisos
    for mes in sorted(ap["meses"], key=chave_mes):
        soma19 = sum(v.get(mes, 0.0) for v in ap["cfops_saida"].values())
        esperado = ap["oper_saidas_s1"].get(mes)
        if esperado is None:
            continue
        dif = abs(soma19 - esperado)
        if dif > tolerancia:
            avisos.append(
                "No arquivo %s a Seção 19 divergiu da Seção 1 em %s no mês %s. O arquivo "
                "pode estar truncado. Confira a exportação do relatório antes de usar o "
                "resultado." % (ap["arquivo"] or "enviado", fmt_brl(dif), fmt_mes(mes)))
    return avisos


# =============================================================================
# Identificação da empresa a partir dos próprios relatórios
#
# O CNPJ está dentro do arquivo (Seções 21/22/23 trazem "<14 dígitos> - <UF>"),
# então o analista não precisa digitá-lo. Esta passada é leve e para no primeiro
# acerto de cada arquivo — não monta série nem soma CFOP.
# =============================================================================
_RE_CNPJ_LINHA = re.compile(r"^(\d{14})\s*-\s*([A-Za-z]{2})?")


def escolher_matriz(estabs: list[dict]):
    """A matriz (ordem 0001) representa a empresa no topo da tela e no Excel.

    Sem a matriz no lote, vale o estabelecimento de menor ordem presente — a razão
    social é a mesma, e é melhor identificar pelo que existe do que construir um
    CNPJ que não veio em arquivo nenhum.
    """
    if not estabs:
        return None
    matriz = [e for e in estabs if e["cnpj"][8:12] == "0001"]
    if matriz:
        return matriz[0]
    return sorted(estabs, key=lambda e: e["cnpj"][8:12])[0]


def identificar_estabelecimentos(arquivos, progress=None) -> dict:
    """Descobre de que estabelecimentos são os relatórios, sem processá-los."""
    def _p(pct, msg):
        if progress:
            progress(int(pct), msg)

    partes = listar_partes(arquivos)
    total = len(partes) or 1
    achados: list[dict] = []
    vistos: set[str] = set()
    nao_reconhecidos: list[str] = []

    for i, parte in enumerate(partes, start=1):
        with abrir_linhas(parte) as fluxo:
            amostra = list(itertools.islice(fluxo, 40))
            if not eh_relatorio_apuracao(parte.nome, amostra):
                nao_reconhecidos.append(os.path.basename(parte.nome))
            else:
                cnpj = uf = ""
                for bruta in itertools.chain(amostra, fluxo):
                    m = _RE_CNPJ_LINHA.match(_clean(bruta.split(";")[0]))
                    if m:
                        cnpj, uf = m.group(1), (m.group(2) or "").upper()
                        break               # achou: não precisa ler o resto
                if not cnpj:
                    nao_reconhecidos.append(os.path.basename(parte.nome))
                elif cnpj not in vistos:    # mesma regra do processamento:
                    vistos.add(cnpj)        # duplicado é descartado, não somado
                    achados.append({"arquivo": os.path.basename(parte.nome),
                                    "cnpj": cnpj, "cnpj_fmt": fmt_cnpj(cnpj), "uf": uf})
        _p(60 * i / total, "Identificando a empresa… (%d/%d)" % (i, total))

    achados.sort(key=lambda e: e["cnpj"])
    raizes = sorted({e["cnpj"][:8] for e in achados})
    matriz = escolher_matriz(achados)
    erro = None
    if len(raizes) > 1:
        erro = MSG_RAIZES_DIFERENTES.format(
            raizes=", ".join(fmt_cnpj(r + "000000")[:10] for r in raizes))
    return {
        "estabelecimentos": achados,
        "matriz": matriz,
        "raizes": raizes,
        "erro": erro,
        "nao_reconhecidos": nao_reconhecidos,
    }


# =============================================================================
# Núcleo
# =============================================================================
def processar_saldo_credor(arquivos, cnpj: str, razao_social: str,
                           progress=None) -> dict:
    """
    arquivos     : lista de caminhos ou (nome, bytes) — relatórios de apuração
    cnpj         : CNPJ consultado na tela (14 dígitos)
    razao_social : razão social vinda da consulta do HUB
    progress     : callback(pct:int, msg:str)

    "Não há oportunidade" é conclusão, não falha: os desfechos 1 e 2 voltam neste
    mesmo dicionário, com status e explicação.
    """
    def _p(pct, msg):
        if progress:
            progress(int(pct), msg)

    _p(3, "Lendo os relatórios de apuração…")
    partes = listar_partes(arquivos)
    total_partes = len(partes) or 1

    apuracoes: list[dict] = []
    vistos: dict[str, str] = {}          # cnpj -> arquivo que ficou
    nao_reconhecidos: list[str] = []
    avisos: list[str] = []

    for i, parte in enumerate(partes, start=1):
        with abrir_linhas(parte) as fluxo:
            amostra = list(itertools.islice(fluxo, 40))
            if not eh_relatorio_apuracao(parte.nome, amostra):
                nao_reconhecidos.append(os.path.basename(parte.nome))
            else:
                ap = parse_apuracao_mensal(itertools.chain(amostra, fluxo), parte.nome)
                anterior = vistos.get(ap["cnpj"])
                if anterior:
                    # arquivo duplicado é descartado: somar dobraria faturamento e
                    # saldo, e fundir séries parciais seria adivinhação
                    avisos.append(
                        "O estabelecimento %s veio em mais de um arquivo (%s e %s). "
                        "Considerei apenas %s e descartei o outro."
                        % (fmt_cnpj(ap["cnpj"]), anterior, ap["arquivo"], anterior))
                else:
                    vistos[ap["cnpj"]] = ap["arquivo"]
                    apuracoes.append(ap)
        _p(3 + 62 * i / total_partes,
           "Lendo arquivos… (%d/%d)" % (i, total_partes))

    if not apuracoes:
        raise ErroDeNegocio(
            "Nenhum relatório de Apuração de ICMS (ICMSProprio) foi reconhecido nos "
            "arquivos enviados. Envie o relatório exportado em CSV, um por "
            "estabelecimento — pode ser em .zip.")

    # Janela não prescrita: recorta a série ANTES de avaliar, para que mês prescrito
    # não apareça em canto nenhum do resultado.
    janela = meses_da_janela([m for ap in apuracoes for m in ap["meses"]])
    fora_da_janela = []
    for ap in apuracoes:
        ap["meses"] = [m for m in ap["meses"] if fmt_mes(m) in janela]
    for ap in [a for a in apuracoes if not a["meses"]]:
        fora_da_janela.append(ap)
        avisos.append(
            "O estabelecimento %s não tem escrituração nos últimos %d meses e ficou fora "
            "da análise — o período dele já está prescrito para recuperação de crédito."
            % (fmt_cnpj(ap["cnpj"]), MESES_ANALISE))
    apuracoes = [ap for ap in apuracoes if ap["meses"]]
    if not apuracoes:
        raise ErroDeNegocio(
            "Nenhum dos relatórios enviados tem escrituração nos últimos %d meses, que é "
            "o prazo não prescrito para recuperação de crédito." % MESES_ANALISE)

    _p(70, "Conferindo a consistência dos arquivos…")
    estabs = [avaliar_estabelecimento(ap) for ap in apuracoes]
    estabs.sort(key=lambda e: e["cnpj"])
    validar_cnpj_raiz(estabs, cnpj, razao_social)
    for ap in apuracoes:
        avisos.extend(conferir_secao1_x_secao19(ap))

    _p(85, "Avaliando o último mês de apuração de cada estabelecimento…")
    calculados = [e for e in estabs if e["status"] == STATUS_CALCULADO]
    total = sum(e["correcao"] for e in calculados)

    # o intervalo exibido reflete SOMENTE o que entrou na análise
    todos_meses = sorted({r["mes"] for e in estabs for r in e["serie"]}, key=chave_mes)
    periodo = ("%s a %s (%d meses)" % (todos_meses[0], todos_meses[-1], len(todos_meses))
               if todos_meses else "")
    if todos_meses:
        velho = aviso_extracao_antiga(todos_meses[-1])
        if velho:
            avisos.insert(0, velho)
    cfops_devolucao = sorted({c for e in estabs for c in e["cfops_devolucao"]})

    _p(100, "Concluído.")
    return {
        "empresa": {
            "cnpj": _so_digitos(cnpj),
            "cnpj_fmt": fmt_cnpj(cnpj),
            "razao_social": razao_social or "",
        },
        "estabelecimentos": estabs,
        "explicacao_consolidada": explicacao_consolidada(estabs, total),
        "status_consolidado": (STATUS_CALCULADO if calculados else
                               (STATUS_SEM_EXPORTACAO
                                if any(e["status"] == STATUS_SEM_EXPORTACAO for e in estabs)
                                else STATUS_SEM_SALDO_CREDOR)),
        "totais": {
            "correcao": total if calculados else None,
            "n_estabelecimentos": len(estabs),
            "n_calculados": len(calculados),
            "n_partes": len(partes),
            "n_nao_reconhecidos": len(nao_reconhecidos),
        },
        "periodo": {"primeiro": todos_meses[0] if todos_meses else "",
                    "ultimo": todos_meses[-1] if todos_meses else "",
                    "coberto": periodo},
        "cfops_devolucao": cfops_devolucao,
        "tem_export_indireta": any(e["tem_export_indireta"] for e in estabs),
        "avisos": avisos[:20],
        "nao_reconhecidos": nao_reconhecidos[:20],
    }


def montar_ressalvas(resultado: dict) -> list[tuple[str, str]]:
    """As notas do rodapé do Excel — e da seção Ressalvas metodológicas do README.

    Cada uma responde a uma pergunta que o cliente ou o fisco faz; a numeração é
    contígua porque as condicionais entram só quando se aplicam.
    """
    notas = [
        ("Critério do mês de referência", RESSALVA_MES),
        ("Período analisado", RESSALVA_JANELA.format(
            limite=MESES_ANALISE, corte=resultado["periodo"]["ultimo"] or "não identificado")),
        ("Natureza do saldo utilizado", RESSALVA_NATUREZA_SALDO),
        ("Composição do faturamento total", RESSALVA_FATURAMENTO),
    ]
    if resultado.get("tem_export_indireta"):
        notas.append(("Exportação indireta", RESSALVA_EXPORT_INDIRETA))
    if resultado.get("cfops_devolucao"):
        notas.append(("Devolução de exportação", RESSALVA_DEVOLUCAO.format(
            cfops="/".join(resultado["cfops_devolucao"]))))
    notas.append(("Origem dos dados", RESSALVA_ORIGEM.format(
        periodo=resultado["periodo"]["coberto"] or "não identificado",
        corte=resultado["periodo"]["ultimo"] or "não identificado")))
    return [("%d. %s" % (i, titulo), texto)
            for i, (titulo, texto) in enumerate(notas, start=1)]


# =============================================================================
# Excel AUDITÁVEL — 3 abas, tipografia e paleta oficiais, FÓRMULAS REAIS
#
# Modo normal (não write_only) de propósito: o bloco de veredito exige célula
# mesclada com wrap_text, e WriteOnlyWorksheet não tem merge_cells. O volume é
# limitado — um estabelecimento rende ~65 linhas na aba de conferência —, então
# não há o problema de memória que obrigou a irmã a usar write_only.
# =============================================================================
# Paleta oficial do Grupo EFCT (seção 4.1 do team-kit). Não há cor fora desta lista
# na planilha; as descontinuadas (#3B82F6, #1F2937) não aparecem.
PETROLEO = "FF001C26"        # azul petróleo, cor-mãe
OLIVA = "FFB3BC2B"           # verde oliva, accent
CIANO = "FF4DBDF5"           # ciano, info
CREME = "FFFEFEDF"           # creme, fundos suaves
TEXTO = "FF111827"           # texto principal
SECUNDARIO = "FF6B7280"      # texto secundário
BORDA = "FFE5E7EB"           # borda sutil
BRANCO = "FFFFFFFF"

# Tipografia oficial (seção 4.2 do team-kit): Bebas Neue nos títulos, Exo 2 no corpo
# e nos labels, Libre Baskerville Italic no texto de ressalva jurídica.
#
# O xlsx não embute fonte: em máquina sem a família instalada o Excel substitui, e o
# arquivo continua legível. Exo 2 cobre quase toda a planilha; Bebas Neue aparece só
# no título das abas e no valor do veredito. Os CORPOS seguem a escala de planilha
# (9-16 pt), não a escala em px da seção 4.2, que é da tela.
FONTE_TITULO = "Bebas Neue"
FONTE_CORPO = "Exo 2"
FONTE_ITALICO = "Libre Baskerville"

# zero renderizado como '—', como manda o padrão de saída
FMT_BRL = r"#,##0.00;\(#,##0.00\);\—"
FMT_PCT = r"0.00%;\(0.00%\);\—"


def gerar_excel_auditavel(resultado: dict, saida_path: str) -> str:
    """Gera o .xlsx. Disponível TAMBÉM quando não houve oportunidade: o relatório
    que documenta a não-aplicabilidade tem valor para o cliente."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.properties import CalcProperties

    def fonte(size=9, bold=False, cor=TEXTO, familia=None, italico=False):
        return Font(name=familia or FONTE_CORPO, size=size, bold=bold, color=cor,
                    italic=italico)

    def fill(cor):
        return PatternFill("solid", fgColor=cor)

    fina = Side(style="thin", color=BORDA)
    grade = Border(left=fina, right=fina, top=fina, bottom=fina)

    esq = Alignment(horizontal="left", vertical="center")
    esq_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ctr = Alignment(horizontal="center", vertical="center")
    ctr_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    dir_ = Alignment(horizontal="right", vertical="center")

    def por(ws, ref, valor=None, f=None, preenche=None, alinha=None, fmt=None,
            borda=False):
        """Escreve no canto superior esquerdo, estende o estilo e mescla a faixa."""
        primeira = ref.split(":")[0]
        cel = ws[primeira]
        if valor is not None:
            cel.value = valor
        if f:
            cel.font = f
        if alinha:
            cel.alignment = alinha
        if fmt:
            cel.number_format = fmt
        if borda:
            cel.border = grade
        if ":" in ref:
            if preenche:
                for linha in ws[ref]:
                    for c in linha:
                        c.fill = fill(preenche)
            ws.merge_cells(ref)
        elif preenche:
            cel.fill = fill(preenche)
        return cel

    estabs = resultado["estabelecimentos"]
    wb = Workbook()
    # openpyxl grava a fórmula, não o valor: sem isto o Excel abriria com célula
    # vazia até o usuário forçar o recálculo
    wb.calculation = CalcProperties(fullCalcOnLoad=True)
    try:
        # fonte padrão da pasta: célula que eu não estilizar já nasce na tipografia
        # oficial. É API interna do openpyxl, então falha aqui não pode derrubar a
        # geração da planilha.
        wb._named_styles["Normal"].font = Font(name=FONTE_CORPO, size=10, color=TEXTO)
    except Exception:  # noqa: BLE001
        pass

    # -------------------------------------------------------------- aba Série
    # Escrita antes da aba Cálculo porque as células calculadas da aba principal
    # apontam para esta: os insumos ficam num só lugar e o revisor segue a
    # referência clicando na célula.
    ws_serie = wb.create_sheet("Série Mensal (conferência)")
    ws_serie.sheet_view.showGridLines = False
    for col, larg in zip("ABCDEFG", (2, 24, 18, 20, 13, 18, 3)):
        ws_serie.column_dimensions[col].width = larg

    por(ws_serie, "B2:F2", "SÉRIE MENSAL — MATERIAL DE CONFERÊNCIA",
        fonte(15, True, BRANCO, FONTE_TITULO), PETROLEO, esq)
    ws_serie.row_dimensions[2].height = 24
    por(ws_serie, "B3:F3",
        "Insumos mês a mês. NÃO integra o cálculo — a correção é apurada "
        "exclusivamente sobre o último mês de apuração, destacado em cada bloco.",
        fonte(9, False, SECUNDARIO), None, esq)

    linha_ref: dict[str, int] = {}       # cnpj -> linha do mês de referência
    r = 5
    for est in estabs:
        por(ws_serie, "B%d:F%d" % (r, r), "%s — %s" % (est["cnpj_fmt"], est["uf"]),
            fonte(10, True, BRANCO), CIANO, esq)
        r += 1
        for j, titulo in enumerate(["Mês", "Faturamento total", "Faturamento de exportação",
                                    "% Exportação", "Saldo credor de ICMS"]):
            c = ws_serie.cell(row=r, column=2 + j, value=titulo)
            c.font, c.fill, c.alignment = fonte(9, True, PETROLEO), fill(CREME), ctr_wrap
            c.border = grade
        r += 1
        primeira_dados = r
        for reg in est["serie"]:
            eh_ref = reg["mes"] == est["mes_ref"]
            rotulo = ("◄ " + reg["mes"]) if eh_ref else reg["mes"]
            destaque = OLIVA if eh_ref else None
            por(ws_serie, "B%d" % r, rotulo, fonte(9, eh_ref, PETROLEO), destaque, ctr, borda=True)
            por(ws_serie, "C%d" % r, round(reg["faturamento"], 2),
                fonte(9, eh_ref), destaque, dir_, FMT_BRL)
            por(ws_serie, "D%d" % r, round(reg["exportacao"], 2),
                fonte(9, eh_ref), destaque, dir_, FMT_BRL)
            por(ws_serie, "E%d" % r, "=IF(C{0}=0,0,D{0}/C{0})".format(r),
                fonte(9, eh_ref), destaque, dir_, FMT_PCT)
            # célula VAZIA (não zero) quando o mês não tem escrituração
            saldo = reg["saldo_credor"]
            por(ws_serie, "F%d" % r, round(saldo, 2) if saldo is not None else None,
                fonte(9, eh_ref), destaque, dir_, FMT_BRL)
            if eh_ref:
                linha_ref[est["cnpj"]] = r
            r += 1
        ultima_dados = r - 1
        por(ws_serie, "B%d" % r, "Total do período", fonte(9, True, PETROLEO), None, dir_)
        por(ws_serie, "C%d" % r, "=SUM(C%d:C%d)" % (primeira_dados, ultima_dados),
            fonte(9, True, PETROLEO), None, dir_, FMT_BRL)
        por(ws_serie, "D%d" % r, "=SUM(D%d:D%d)" % (primeira_dados, ultima_dados),
            fonte(9, True, PETROLEO), None, dir_, FMT_BRL)
        # saldo credor é ESTOQUE: somar meses não significaria nada
        por(ws_serie, "F%d" % r, "—", fonte(9, False, SECUNDARIO), None, dir_)
        r += 3
    por(ws_serie, "B%d:F%d" % (r, r),
        "Célula de saldo credor vazia indica mês sem escrituração do estabelecimento "
        "(não é zero). O total do período não se aplica ao saldo credor, por ser "
        "grandeza de estoque e não de fluxo.",
        fonte(9, False, SECUNDARIO), None, esq)

    # ------------------------------------------------------------ aba Cálculo
    ws = wb.create_sheet("Cálculo", 0)
    ws.sheet_view.showGridLines = False
    # B com 28: os rótulos da identificação ("Estabelecimentos analisados", "Mês de
    # corte da extração") ficavam CORTADOS em 20, porque a célula ao lado tem valor e
    # não deixa o texto transbordar. Documento que pode virar anexo processual não
    # pode sair com rótulo truncado.
    for col, larg in zip("ABCDEFGHIJK",
                         (2, 28, 6, 14, 17, 17, 12, 17, 26, 17, 95)):
        ws.column_dimensions[col].width = larg

    por(ws, "B2:K2", "CORREÇÃO DO SALDO CREDOR DE ICMS PROPORCIONAL ÀS EXPORTAÇÕES",
        fonte(18, True, BRANCO, FONTE_TITULO), PETROLEO, esq)
    ws.row_dimensions[2].height = 28

    identificacao = [
        ("Empresa", resultado["empresa"]["razao_social"] or "—"),
        ("CNPJ", resultado["empresa"]["cnpj_fmt"]),
        ("Estabelecimentos analisados", len(estabs)),
        # "analisado", não "coberto": os arquivos podem trazer mais meses do que a
        # janela não prescrita, e o que vale para o leitor é o que entrou na conta
        ("Período analisado", resultado["periodo"]["coberto"] or "—"),
        ("Mês de corte da extração", resultado["periodo"]["ultimo"] or "—"),
    ]
    r = 4
    for rotulo, valor in identificacao:
        por(ws, "B%d" % r, rotulo, fonte(10, True, PETROLEO), None, esq)
        por(ws, "C%d:F%d" % (r, r), valor, fonte(10), None, esq)
        ws.row_dimensions[r].height = 15
        r += 1

    por(ws, "B10:K10", "RESULTADO DA ANÁLISE", fonte(11, True, PETROLEO), CREME, esq)
    total = resultado["totais"]["correcao"]
    # sem correção calculada: NENHUM valor monetário aqui. Um "R$ 0,00" seria
    # lido como "calculamos e deu zero", que é afirmação diferente e errada.
    por(ws, "B11:D12", round(total, 2) if total is not None else "SEM OPORTUNIDADE",
        fonte(18, True, BRANCO, FONTE_TITULO), PETROLEO, ctr_wrap,
        FMT_BRL if total is not None else None)
    por(ws, "E11:K12", resultado["explicacao_consolidada"], fonte(10), CREME, esq_wrap)
    ws.row_dimensions[11].height = 30
    ws.row_dimensions[12].height = 30

    cabecalhos = ["CNPJ", "UF", "Último mês de apuração", "Faturamento total do mês",
                  "Faturamento de exportação do mês", "% Exportação",
                  "Saldo credor de ICMS", "Status", "Correção (R$)", "Explicação"]
    for j, titulo in enumerate(cabecalhos):
        c = ws.cell(row=14, column=2 + j, value=titulo)
        c.font, c.fill, c.alignment = fonte(9, True, BRANCO), fill(PETROLEO), ctr_wrap
        c.border = grade
    ws.row_dimensions[14].height = 31.5

    SERIE = "'Série Mensal (conferência)'"
    r = 15
    primeira = r
    for est in estabs:
        ref = linha_ref.get(est["cnpj"])
        por(ws, "B%d" % r, est["cnpj_fmt"], fonte(9, False, TEXTO), None, ctr, borda=True)
        por(ws, "C%d" % r, est["uf"], fonte(9), None, ctr, borda=True)
        # o mês de referência é o eixo do cálculo: destacado, como no modelo
        por(ws, "D%d" % r, est["mes_ref"], fonte(9, True, PETROLEO), None, ctr, borda=True)
        # os insumos apontam para a linha do mês de referência na aba de
        # conferência — uma só fonte de verdade, conferível clicando na célula
        por(ws, "E%d" % r,
            "=%s!C%d" % (SERIE, ref) if ref else round(est["faturamento"], 2),
            fonte(9), None, dir_, FMT_BRL, borda=True)
        por(ws, "F%d" % r,
            "=%s!D%d" % (SERIE, ref) if ref else round(est["exportacao"], 2),
            fonte(9), None, dir_, FMT_BRL, borda=True)
        por(ws, "G%d" % r,
            '=IF($I{0}<>"Calculado","",IF(E{0}=0,0,F{0}/E{0}))'.format(r),
            fonte(9, True), None, dir_, FMT_PCT, borda=True)
        por(ws, "H%d" % r,
            "=%s!F%d" % (SERIE, ref) if ref else round(est["saldo_credor"], 2),
            fonte(9), None, dir_, FMT_BRL, borda=True)
        por(ws, "I%d" % r, est["rotulo_status"], fonte(9, True, SECUNDARIO), None, ctr_wrap,
            borda=True)
        por(ws, "J%d" % r, '=IF($I{0}<>"Calculado","",H{0}*G{0})'.format(r),
            fonte(10, True, PETROLEO), None, dir_, FMT_BRL, borda=True)
        por(ws, "K%d" % r, est["explicacao"], fonte(9, False, SECUNDARIO), None, esq_wrap,
            borda=True)
        ws.row_dimensions[r].height = 45.75
        r += 1
    ultima = r - 1

    for j in range(2, 12):
        c = ws.cell(row=r, column=j)
        c.fill = fill(PETROLEO)
        c.font = fonte(10, True, BRANCO)
        c.alignment = esq
    por(ws, "B%d" % r, "TOTAL", fonte(10, True, BRANCO), PETROLEO, esq)
    por(ws, "I%d" % r, "%d de %d qualificaram"
        % (resultado["totais"]["n_calculados"], len(estabs)),
        fonte(9, True, BRANCO), PETROLEO, ctr)
    # as células não calculadas ficam em branco, então SUM já soma só o que foi
    # calculado; COUNT evita exibir 0,00 quando nada qualificou
    por(ws, "J%d" % r,
        '=IF(COUNT(J{0}:J{1})=0,"",SUM(J{0}:J{1}))'.format(primeira, ultima),
        fonte(12, True, BRANCO), PETROLEO, dir_, FMT_BRL)
    r += 2

    por(ws, "B%d:K%d" % (r, r), "RESSALVAS", fonte(10, True, BRANCO), PETROLEO, esq)
    r += 1
    for titulo, texto in montar_ressalvas(resultado):
        # a nota que distingue saldo a transportar de crédito acumulado é a de peso
        # jurídico: vai no itálico elegante do padrão (Libre Baskerville Italic)
        juridica = "Natureza do saldo" in titulo
        corpo = (fonte(9, False, SECUNDARIO, FONTE_ITALICO, italico=True) if juridica
                 else fonte(9, False, SECUNDARIO))
        por(ws, "B%d:C%d" % (r, r + 1), titulo, fonte(9, True, PETROLEO), CREME, esq_wrap)
        por(ws, "D%d:K%d" % (r, r + 1), texto, corpo, CREME, esq_wrap)
        ws.row_dimensions[r].height = 24
        ws.row_dimensions[r + 1].height = 24
        r += 2

    # -------------------------------------------------------- aba Exportações
    ws_exp = wb.create_sheet("Exportações")
    ws_exp.sheet_view.showGridLines = False
    for col, larg in zip("ABCDEFG", (2, 20, 11, 8, 34, 17, 11)):
        ws_exp.column_dimensions[col].width = larg
    por(ws_exp, "B2:G2", "OPERAÇÕES DE EXPORTAÇÃO IDENTIFICADAS",
        fonte(15, True, BRANCO, FONTE_TITULO), PETROLEO, esq)
    ws_exp.row_dimensions[2].height = 24
    por(ws_exp, "B3:G3",
        "Prova documental da existência e do volume das exportações no período. "
        "Linhas do último mês de apuração destacadas.",
        fonte(9, False, SECUNDARIO), None, esq)

    for j, titulo in enumerate(["CNPJ", "Mês", "CFOP", "Descrição do CFOP",
                                "Valor da operação (R$)", "Tipo"]):
        c = ws_exp.cell(row=5, column=2 + j, value=titulo)
        c.font, c.fill, c.alignment = fonte(9, True, BRANCO), fill(PETROLEO), ctr_wrap
        c.border = grade
    def linha_operacao(est, op, r):
        eh_ref = op["mes"] == est["mes_ref"]
        destaque = OLIVA if eh_ref else None
        por(ws_exp, "B%d" % r, est["cnpj_fmt"], fonte(9, eh_ref), destaque, ctr, borda=True)
        por(ws_exp, "C%d" % r, op["mes"], fonte(9, eh_ref), destaque, ctr, borda=True)
        por(ws_exp, "D%d" % r, op["cfop"], fonte(9, eh_ref), destaque, ctr, borda=True)
        por(ws_exp, "E%d" % r, op["descricao"], fonte(9, eh_ref), destaque, esq, borda=True)
        por(ws_exp, "F%d" % r, round(op["valor"], 2),
            fonte(9, eh_ref), destaque, dir_, FMT_BRL)
        por(ws_exp, "G%d" % r, op["tipo"], fonte(9, eh_ref), destaque, ctr, borda=True)

    def faixa_total(r, rotulo, formula):
        for j in range(2, 8):
            c = ws_exp.cell(row=r, column=j)
            c.fill, c.font = fill(PETROLEO), fonte(9, True, BRANCO)
        por(ws_exp, "B%d" % r, rotulo, fonte(10, True, BRANCO), PETROLEO, esq)
        por(ws_exp, "F%d" % r, formula, fonte(10, True, BRANCO), PETROLEO, dir_, FMT_BRL)

    r = 6
    primeira_exp = r
    for est in estabs:
        saidas = [o for o in est["exportacoes"] if o["valor"] > 0]
        if not saidas:
            por(ws_exp, "B%d" % r, est["cnpj_fmt"], fonte(9), None, ctr, "@")
            por(ws_exp, "C%d:G%d" % (r, r),
                "Nenhuma operação de exportação identificada no período",
                fonte(9, False, SECUNDARIO), None, esq)
            r += 1
            continue
        for op in saidas:
            linha_operacao(est, op, r)
            r += 1
    devolucoes = [(est, o) for est in estabs for o in est["exportacoes"] if o["valor"] < 0]
    # "TOTAL" puro quando não há bloco de devolução abaixo — é o rótulo do modelo; com
    # o bloco, o total precisa dizer a que se refere
    faixa_total(r, "TOTAL EXPORTADO NO PERÍODO" if devolucoes else "TOTAL",
                "=SUM(F%d:F%d)" % (primeira_exp, r - 1))
    r += 2

    # Devolução/retorno de exportação em bloco PRÓPRIO, fora do total exportado:
    # é ajuste do mês em que ocorreu, não operação de exportação. Somá-la ao total
    # faria a aba divergir da série mensal quando o piso em zero é acionado.
    if devolucoes:
        por(ws_exp, "B%d:G%d" % (r, r),
            "DEVOLUÇÕES / RETORNOS DE EXPORTAÇÃO ABATIDOS DO MÊS DE OCORRÊNCIA",
            fonte(10, True, BRANCO), PETROLEO, esq)
        r += 1
        primeira_dev = r
        for est, op in devolucoes:
            linha_operacao(est, op, r)
            r += 1
        faixa_total(r, "TOTAL DEVOLVIDO", "=SUM(F%d:F%d)" % (primeira_dev, r - 1))
        r += 1
        limitados = [(est["cnpj_fmt"], m) for est in estabs for m in est["meses_limitados"]]
        if limitados:
            por(ws_exp, "B%d:G%d" % (r, r),
                "Nos meses a seguir o valor devolvido supera a exportação do próprio mês, "
                "e a exportação considerada no cálculo foi limitada a zero (nunca "
                "negativa): " + "; ".join("%s em %s" % (c, m) for c, m in limitados) + ".",
                fonte(9, False, SECUNDARIO), None, esq_wrap)
            ws_exp.row_dimensions[r].height = 28

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    destino = os.path.dirname(os.path.abspath(saida_path))
    if destino:
        os.makedirs(destino, exist_ok=True)
    wb.save(saida_path)
    return saida_path
